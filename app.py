#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Radar JF — Interface Web para o Monitor de Ações DJEN/CNJ
=======================================================
Servidor FastAPI que expõe o monitoramento via interface web moderna.

Uso:
    pip install -r requirements.txt
    uvicorn app:app --reload
"""

import os
import json
import asyncio
import uuid
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Query, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.security import APIKeyHeader

# Importa a lógica do monitor existente
import monitor_djen as monitor

load_dotenv()

AUTH_KEY = os.getenv("AUTH_KEY", "")
if not AUTH_KEY:
    print("⚠  AUTH_KEY não definida no .env — crie um arquivo .env com AUTH_KEY=sua-chave")

app = FastAPI(title="Radar JF — Monitor DJEN/CNJ")

# ---------------------------------------------------------------------------
# Diretórios
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
UPLOADS_DIR = BASE_DIR / "uploads"
RESULTS_DIR = BASE_DIR / "results"
UPLOADS_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Estado da aplicação (em memória)
# ---------------------------------------------------------------------------
# Arquivo base ativo para monitoramento
active_file: Optional[str] = None
# Jobs de varredura em andamento / concluídos
jobs: dict = {}

# ---------------------------------------------------------------------------
# Autenticação simples por API Key
# ---------------------------------------------------------------------------
api_key_header = APIKeyHeader(name="X-Auth-Key", auto_error=False)


async def verify_auth(request: Request, key: Optional[str] = Depends(api_key_header)):
    """Verifica a chave de autenticação via header ou cookie."""
    if not AUTH_KEY:
        return True  # sem chave configurada, aceita tudo (dev mode)
    # Tenta header primeiro, depois cookie
    token = key or request.cookies.get("auth_key")
    if token != AUTH_KEY:
        raise HTTPException(status_code=401, detail="Chave de autenticação inválida")
    return True


# ---------------------------------------------------------------------------
# Rotas — Páginas
# ---------------------------------------------------------------------------
@app.get("/")
async def root():
    html_path = BASE_DIR / "templates" / "index.html"
    print(f"[DEBUG] BASE_DIR={BASE_DIR}, html_path={html_path}, exists={html_path.exists()}")
    if not html_path.exists():
        return JSONResponse({"detail": f"Template not found: {html_path}"}, status_code=500)
    return HTMLResponse(content=html_path.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# Rotas — Auth
# ---------------------------------------------------------------------------
@app.post("/api/auth/login")
async def login(body: dict):
    key = body.get("key", "")
    if not AUTH_KEY:
        return {"ok": True, "message": "Sem autenticação configurada"}
    if key != AUTH_KEY:
        raise HTTPException(status_code=401, detail="Chave inválida")
    response = JSONResponse({"ok": True, "message": "Autenticado"})
    response.set_cookie("auth_key", key, httponly=True, samesite="strict", max_age=86400 * 7)
    return response


@app.post("/api/auth/logout")
async def logout():
    response = JSONResponse({"ok": True})
    response.delete_cookie("auth_key")
    return response


@app.get("/api/auth/check")
async def check_auth(request: Request):
    if not AUTH_KEY:
        return {"authenticated": True}
    token = request.cookies.get("auth_key")
    return {"authenticated": token == AUTH_KEY}


# ---------------------------------------------------------------------------
# Rotas — Arquivo base
# ---------------------------------------------------------------------------
@app.post("/api/files/upload")
async def upload_file(file: UploadFile = File(...), _=Depends(verify_auth)):
    global active_file
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Apenas arquivos Excel (.xlsx) são aceitos")

    dest = UPLOADS_DIR / file.filename
    content = await file.read()
    dest.write_bytes(content)
    active_file = str(dest)
    return {"ok": True, "filename": file.filename, "path": str(dest)}


@app.get("/api/files/list")
async def list_files(_=Depends(verify_auth)):
    files = []
    for f in UPLOADS_DIR.glob("*.xlsx"):
        files.append({"name": f.name, "size": f.stat().st_size, "active": str(f) == active_file})
    # Inclui arquivo padrão se existir na raiz
    default = BASE_DIR / monitor.PLANILHA_ENTRADA
    if default.exists() and not any(f["name"] == default.name for f in files):
        files.insert(0, {"name": default.name, "size": default.stat().st_size,
                         "active": str(default) == active_file, "default": True})
    return {"files": files}


@app.post("/api/files/select")
async def select_file(body: dict, _=Depends(verify_auth)):
    global active_file
    name = body.get("filename", "")
    # Procura em uploads
    path = UPLOADS_DIR / name
    if not path.exists():
        # Tenta na raiz (arquivo padrão)
        path = BASE_DIR / name
    if not path.exists():
        raise HTTPException(404, f"Arquivo '{name}' não encontrado")
    active_file = str(path)
    return {"ok": True, "filename": name}


# ---------------------------------------------------------------------------
# Rotas — Monitoramento
# ---------------------------------------------------------------------------
@app.post("/api/monitor/start")
async def start_monitor(body: dict, _=Depends(verify_auth)):
    global active_file

    dias = body.get("dias", 30)
    if dias not in [7, 30, 90]:
        raise HTTPException(400, "Dias deve ser 7, 30 ou 90")

    tribunais = body.get("tribunais", monitor.TRIBUNAIS)
    varrer_inst = body.get("instituicoes", True)
    varrer_sind = body.get("sindicatos", True)
    varrer_esc = body.get("escritorios", True)

    # Determina arquivo base
    planilha = active_file
    if not planilha:
        default = BASE_DIR / monitor.PLANILHA_ENTRADA
        if default.exists():
            planilha = str(default)
        else:
            raise HTTPException(400, "Nenhum arquivo base selecionado. Faça upload de uma planilha.")

    if not Path(planilha).exists():
        raise HTTPException(400, f"Arquivo '{planilha}' não encontrado")

    job_id = str(uuid.uuid4())[:8]
    jobs[job_id] = {"status": "running", "progress": 0, "total": 0, "results": [], "log": []}

    # Roda a varredura em background
    asyncio.create_task(_run_monitor(job_id, planilha, dias, tribunais,
                                     varrer_inst, varrer_sind, varrer_esc))
    return {"ok": True, "job_id": job_id}


async def _run_monitor(job_id, planilha, dias, tribunais, varrer_inst, varrer_sind, varrer_esc):
    """Executa a varredura em background, atualizando o job."""
    job = jobs[job_id]
    try:
        fim = date.today()
        ini = fim - timedelta(days=dias)
        ini_s, fim_s = ini.isoformat(), fim.isoformat()

        # Carrega alvos
        inst, sind, esc = monitor.carregar_alvos(planilha, monitor.ABA_ENTRADA)
        job["log"].append(f"Base: {len(inst)} instituições | {len(sind)} sindicatos | {len(esc)} escritórios")
        job["log"].append(f"Janela: {ini_s} a {fim_s} ({dias} dias) | Tribunais: {', '.join(tribunais)}")

        alvos = []
        if varrer_inst:
            alvos += [("nome", a) for a in inst]
        if varrer_sind:
            alvos += [("nome", a) for a in sind]
        if varrer_esc:
            alvos += [("oab", a) for a in esc]

        job["total"] = len(alvos)
        vistos = set()
        resultados = []

        for i, (modo, alvo) in enumerate(alvos, 1):
            job["progress"] = i
            job["current"] = alvo["nome"]
            achados = 0

            for trib in tribunais:
                if modo == "nome":
                    itens = await asyncio.to_thread(
                        monitor.consultar, trib, ini_s, fim_s, nome_parte=alvo["nome_busca"]
                    )
                else:
                    oab = alvo.get("oab_soc") or alvo.get("oab_adv")
                    if not oab:
                        continue
                    num, uf = monitor._split_oab(oab)
                    itens = await asyncio.to_thread(
                        monitor.consultar, trib, ini_s, fim_s, oab=num, uf_oab=uf
                    )

                for it in itens:
                    chave = it.get("id")
                    chave_par = (alvo["nome"], chave)
                    if chave_par in vistos:
                        continue
                    vistos.add(chave_par)
                    resultados.append(monitor.linha_saida(alvo, it))
                    achados += 1

            status_msg = f"[{i}/{len(alvos)}] {alvo['tipo']}: {alvo['nome']} — {achados} comunicação(ões)"
            job["log"].append(status_msg)

        job["results"] = resultados
        job["status"] = "done"
        job["log"].append(f"\n✔ {len(resultados)} comunicação(ões) encontradas")

        # Salva resultado em Excel
        result_file = RESULTS_DIR / f"resultado_{job_id}.xlsx"
        _save_results(resultados, (ini_s, fim_s), str(result_file))
        job["result_file"] = result_file.name

    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        job["log"].append(f"ERRO: {e}")


def _save_results(resultados, periodo, path):
    """Salva resultados em Excel (reutiliza a lógica do monitor)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    for c, nome in enumerate(monitor.COLUNAS, 1):
        cell = ws.cell(1, c, nome)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = monitor.LARGURAS[c - 1]

    fonte = Font(name="Arial", size=10)
    for linha in resultados:
        ws.append([linha.get(k, "") for k in monitor.COLUNAS])
        for c in range(1, len(monitor.COLUNAS) + 1):
            ws.cell(ws.max_row, c).font = fonte
            ws.cell(ws.max_row, c).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(monitor.COLUNAS))}{ws.max_row}"
    wb.save(path)


@app.get("/api/monitor/status/{job_id}")
async def job_status(job_id: str, _=Depends(verify_auth)):
    if job_id not in jobs:
        raise HTTPException(404, "Job não encontrado")
    job = jobs[job_id]
    return {
        "status": job["status"],
        "progress": job.get("progress", 0),
        "total": job.get("total", 0),
        "current": job.get("current", ""),
        "log": job.get("log", [])[-20:],  # últimas 20 linhas
        "result_count": len(job.get("results", [])),
        "error": job.get("error"),
        "result_file": job.get("result_file"),
    }


@app.get("/api/monitor/results/{job_id}")
async def job_results(job_id: str, _=Depends(verify_auth)):
    if job_id not in jobs:
        raise HTTPException(404, "Job não encontrado")
    job = jobs[job_id]
    if job["status"] != "done":
        raise HTTPException(400, "Job ainda não concluído")
    return {"results": job["results"]}


@app.get("/api/monitor/download/{job_id}")
async def download_result(job_id: str, _=Depends(verify_auth)):
    if job_id not in jobs:
        raise HTTPException(404, "Job não encontrado")
    job = jobs[job_id]
    if not job.get("result_file"):
        raise HTTPException(400, "Resultado não disponível")
    path = RESULTS_DIR / job["result_file"]
    if not path.exists():
        raise HTTPException(404, "Arquivo não encontrado")
    return FileResponse(path, filename=f"Resultado_Monitoramento_{job_id}.xlsx",
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 3333))
    uvicorn.run("app:app", host="0.0.0.0", port=port, reload=os.getenv("ENV") != "production")
