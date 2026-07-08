#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Monitor de Ações — DJEN/Comunica (CNJ)
=======================================
Varre o Diário de Justiça Eletrônico Nacional (DJEN) buscando comunicações
(intimações, despachos, decisões, sentenças) em que os ALVOS da planilha-base
aparecem como parte — instituições federais e sindicatos (por NOME) e
escritórios (por OAB).

Fonte 100% pública e gratuita: https://comunicaapi.pje.jus.br/api/v1/comunicacao
(sem gov.br, sem certificado, sem CAPTCHA).

COMO USAR (no seu computador / servidor):
    1) Instale as dependências uma única vez:
         pip install requests openpyxl
    2) Deixe a planilha-base na mesma pasta (ou ajuste PLANILHA_ENTRADA).
    3) Rode:
         python monitor_djen.py                 # usa a janela padrão (JANELA_DIAS)
         python monitor_djen.py --dias 7        # últimos 7 dias
         python monitor_djen.py --dias 30       # últimos 30 dias
         python monitor_djen.py --dias 90       # últimos 90 dias
    4) O resultado sai em: Resultado_Monitoramento_DJEN.xlsx

Nada aqui envia dados a lugar nenhum além da API oficial do CNJ (só leitura).
"""

import argparse
import sys
import time
from datetime import date, timedelta

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =====================================================================
# CONFIGURAÇÃO  (mexa aqui)
# =====================================================================

# Janela de tempo padrão em dias. Pode ser trocada na linha de comando (--dias).
# Valores pensados para o filtro: 7, 30 ou 90.
JANELA_DIAS = 30

# Tribunais a varrer. Padrão = os 6 TRFs federais (foco: servidor público federal).
# Para incluir Justiça estadual, acrescente ex.: "TJBA", "TJSP", "TJRJ".
TRIBUNAIS = ["TRF1", "TRF2", "TRF3", "TRF4", "TRF5", "TRF6"]

# Quais blocos da planilha varrer (True/False).
VARRER_INSTITUICOES = True    # 110 universidades/IFs/CEFETs/Colégio Pedro II (polo passivo, por nome)
VARRER_SINDICATOS   = True    # sindicatos (por nome / razão social)
VARRER_ESCRITORIOS  = True    # escritórios (por OAB) — só roda onde a OAB estiver preenchida

# Arquivos
PLANILHA_ENTRADA = "Monitoramento_de_Acoes__2_.xlsx"
ABA_ENTRADA      = "Monitoramento"
PLANILHA_SAIDA   = "Resultado_Monitoramento_DJEN.xlsx"

# Educação com o servidor do CNJ: pausa entre chamadas e tentativas em caso de falha.
PAUSA_SEGUNDOS   = 0.2
TENTATIVAS       = 3
ITENS_POR_PAGINA = 100
MAX_PAGINAS      = 20          # trava de segurança por alvo/tribunal

API_URL = "https://comunicaapi.pje.jus.br/api/v1/comunicacao"

# =====================================================================
# LEITURA DA PLANILHA-BASE
# =====================================================================

# Cabeçalhos de seção que separam os blocos na coluna A.
_MARCADORES_INSTITUICAO = (
    "UNIVERSIDADES FEDERAIS",
    "INSTITUTOS FEDERAIS",
    "CEFET",
)
# Prefixos típicos de sindicato/associação (para separar de escritório).
_PREFIXOS_SINDICATO = (
    "SIND", "SINT", "ASS", "ASUF", "ASAV", "ASUN", "APTA", "APUFP",
    "FASUBRA", "ANDES", "SINASEFE", "STU", "SISTA", "ASSOCIA",
)


def _limpa_nome(txt):
    """Remove parênteses (acrônimos) e espaços para casar melhor no DJEN."""
    if not txt:
        return ""
    t = str(txt).strip()
    if "(" in t:
        t = t.split("(")[0].strip()
    return " ".join(t.split())


def carregar_alvos(caminho, aba):
    wb = load_workbook(caminho, data_only=True)
    ws = wb[aba]

    # Descobre as colunas pelo cabeçalho (linha 2), com defaults seguros.
    cab = {str(ws.cell(2, c).value).strip().lower(): c
           for c in range(1, ws.max_column + 1) if ws.cell(2, c).value}
    col_nome    = 1
    col_cnpj    = cab.get("cnpj", 6)
    col_oab_soc = cab.get("oab sociedade", 7)
    col_oab_adv = cab.get("oab advogado responsável", 8)
    # Coluna OPCIONAL: se você criar uma coluna chamada "Nome de busca (DJEN)"
    # e preencher a razão social de um sindicato, o script usa ela no lugar da sigla.
    col_busca   = cab.get("nome de busca (djen)")

    instituicoes, sindicatos, escritorios = [], [], []
    secao = "escritorio"

    for r in range(3, ws.max_row + 1):
        bruto = ws.cell(r, col_nome).value
        nome = (str(bruto).strip() if bruto else "")
        if not nome:
            continue
        up = nome.upper()

        if any(up.startswith(m) for m in _MARCADORES_INSTITUICAO):
            secao = "instituicao"
            continue

        cnpj = ws.cell(r, col_cnpj).value
        oab_soc = ws.cell(r, col_oab_soc).value if col_oab_soc else None
        oab_adv = ws.cell(r, col_oab_adv).value if col_oab_adv else None
        busca = ws.cell(r, col_busca).value if col_busca else None

        registro = {
            "linha": r,
            "nome": nome,
            "nome_busca": _limpa_nome(busca) or _limpa_nome(nome),
            "cnpj": cnpj,
            "oab_soc": (str(oab_soc).strip() if oab_soc else None),
            "oab_adv": (str(oab_adv).strip() if oab_adv else None),
        }

        if secao == "instituicao":
            registro["tipo"] = "Instituição"
            instituicoes.append(registro)
        elif cnpj or any(up.startswith(p) for p in _PREFIXOS_SINDICATO):
            registro["tipo"] = "Sindicato"
            sindicatos.append(registro)
        else:
            registro["tipo"] = "Escritório"
            escritorios.append(registro)

    return instituicoes, sindicatos, escritorios


# =====================================================================
# CONSULTA AO DJEN
# =====================================================================

def _get(params):
    """GET com tentativas. Retorna dict JSON ou None (ex.: OAB inexistente => 500)."""
    for tentativa in range(1, TENTATIVAS + 1):
        try:
            resp = requests.get(API_URL, params=params, timeout=40)
            if resp.status_code == 200:
                return resp.json()
            # 500 costuma ser "sem resultado" para OAB inexistente — não insiste.
            if resp.status_code in (400, 404, 500):
                return None
            time.sleep(PAUSA_SEGUNDOS * tentativa)
        except requests.RequestException:
            time.sleep(PAUSA_SEGUNDOS * tentativa)
    return None


def consultar(tribunal, ini, fim, nome_parte=None, oab=None, uf_oab=None):
    """Consulta paginada. Devolve lista de itens (comunicações)."""
    base = {
        "siglaTribunal": tribunal,
        "dataDisponibilizacaoInicio": ini,
        "dataDisponibilizacaoFim": fim,
        "itensPorPagina": ITENS_POR_PAGINA,
    }
    if nome_parte:
        base["nomeParte"] = nome_parte
    if oab:
        base["numeroOab"] = oab
        if uf_oab:
            base["ufOab"] = uf_oab

    itens, pagina = [], 1
    while pagina <= MAX_PAGINAS:
        params = dict(base, pagina=pagina)
        data = _get(params)
        time.sleep(PAUSA_SEGUNDOS)
        if not data or not data.get("items"):
            break
        lote = data["items"]
        itens.extend(lote)
        total = data.get("count", len(itens))
        if len(itens) >= total or len(lote) < ITENS_POR_PAGINA:
            break
        pagina += 1
    return itens


# =====================================================================
# NORMALIZAÇÃO DOS RESULTADOS
# =====================================================================

def _resumo_partes(item):
    linhas = []
    for d in (item.get("destinatarios") or []):
        polo = {"A": "Autor", "P": "Réu"}.get(d.get("polo"), d.get("polo") or "")
        linhas.append(f"{polo}: {d.get('nome','')}")
    return " | ".join(linhas)


def _resumo_advs(item):
    out = []
    for a in (item.get("destinatarioadvogados") or []):
        adv = a.get("advogado") or {}
        oab = f"{adv.get('numero_oab','')}/{adv.get('uf_oab','')}".strip("/")
        nome = adv.get("nome", "")
        out.append(f"{nome} (OAB {oab})" if oab else nome)
    return " | ".join(out)


def _polo_do_alvo(item, nome_alvo):
    alvo = (nome_alvo or "").upper()
    for d in (item.get("destinatarios") or []):
        if alvo and alvo in (d.get("nome") or "").upper():
            return {"A": "Autor", "P": "Réu"}.get(d.get("polo"), d.get("polo") or "")
    return ""


def linha_saida(alvo, item):
    texto = (item.get("texto") or "").replace("\n", " ").strip()
    resumo = (texto[:600] + "…") if len(texto) > 600 else texto
    return {
        "Alvo": alvo["nome"],
        "Tipo do alvo": alvo["tipo"],
        "Tribunal": item.get("siglaTribunal", ""),
        "Nº do processo": item.get("numeroprocessocommascara") or item.get("numero_processo", ""),
        "Classe": item.get("nomeClasse", ""),
        "Órgão julgador": item.get("nomeOrgao", ""),
        "Tipo de comunicação": item.get("tipoComunicacao", ""),
        "Documento": item.get("tipoDocumento", ""),
        "Data disponibilização": item.get("datadisponibilizacao") or item.get("data_disponibilizacao", ""),
        "Polo do alvo": _polo_do_alvo(item, alvo["nome_busca"]) or _polo_do_alvo(item, alvo["nome"]),
        "Partes": _resumo_partes(item),
        "Advogados (OAB)": _resumo_advs(item),
        "Resumo do teor": resumo,
        "Link do processo": item.get("link", ""),
        "ID comunicação": item.get("id", ""),
    }


# =====================================================================
# VARREDURA
# =====================================================================

def varrer(dias):
    fim = date.today()
    ini = fim - timedelta(days=dias)
    ini_s, fim_s = ini.isoformat(), fim.isoformat()

    inst, sind, esc = carregar_alvos(PLANILHA_ENTRADA, ABA_ENTRADA)
    print(f"Base: {len(inst)} instituições | {len(sind)} sindicatos | {len(esc)} escritórios")
    print(f"Janela: {ini_s} a {fim_s} ({dias} dias) | Tribunais: {', '.join(TRIBUNAIS)}\n")

    alvos = []
    if VARRER_INSTITUICOES:
        alvos += [("nome", a) for a in inst]
    if VARRER_SINDICATOS:
        alvos += [("nome", a) for a in sind]
    if VARRER_ESCRITORIOS:
        alvos += [("oab", a) for a in esc]

    vistos = set()      # (id_comunicacao) para não duplicar
    resultados = []
    sem_oab = 0

    for i, (modo, alvo) in enumerate(alvos, 1):
        rotulo = f"[{i}/{len(alvos)}] {alvo['tipo']}: {alvo['nome']}"
        achados = 0
        for trib in TRIBUNAIS:
            if modo == "nome":
                itens = consultar(trib, ini_s, fim_s, nome_parte=alvo["nome_busca"])
            else:  # oab
                oab = alvo.get("oab_soc") or alvo.get("oab_adv")
                if not oab:
                    continue
                num, uf = _split_oab(oab)
                itens = consultar(trib, ini_s, fim_s, oab=num, uf_oab=uf)

            for it in itens:
                chave = it.get("id")
                # dedup por comunicação; mantém quando o mesmo processo bate em alvos diferentes
                chave_par = (alvo["nome"], chave)
                if chave_par in vistos:
                    continue
                vistos.add(chave_par)
                resultados.append(linha_saida(alvo, it))
                achados += 1

        if modo == "oab" and not (alvo.get("oab_soc") or alvo.get("oab_adv")):
            sem_oab += 1
            print(f"{rotulo}  —  (sem OAB preenchida; pulado)")
        else:
            print(f"{rotulo}  —  {achados} comunicação(ões)")

    if sem_oab:
        print(f"\n⚠  {sem_oab} escritório(s) sem OAB na planilha foram pulados. "
              f"Preencha as colunas 'OAB Sociedade' / 'OAB Advogado responsável' para incluí-los.")
    return resultados, (ini_s, fim_s)


def _split_oab(valor):
    """Aceita '12345/BA', '12345 BA' ou '12345'. Retorna (numero, uf|None)."""
    v = str(valor).upper().replace("-", " ").replace("/", " ").split()
    num = "".join(ch for ch in v[0] if ch.isdigit()) if v else ""
    uf = next((p for p in v if len(p) == 2 and p.isalpha()), None)
    return num, uf


# =====================================================================
# SAÍDA EM EXCEL
# =====================================================================

COLUNAS = ["Alvo", "Tipo do alvo", "Tribunal", "Nº do processo", "Classe",
           "Órgão julgador", "Tipo de comunicação", "Documento",
           "Data disponibilização", "Polo do alvo", "Partes",
           "Advogados (OAB)", "Resumo do teor", "Link do processo", "ID comunicação"]

LARGURAS = [34, 12, 9, 24, 26, 34, 18, 16, 16, 12, 46, 40, 70, 40, 16]


def salvar(resultados, periodo):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    for c, nome in enumerate(COLUNAS, 1):
        cell = ws.cell(1, c, nome)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = LARGURAS[c - 1]

    fonte = Font(name="Arial", size=10)
    for linha in resultados:
        ws.append([linha.get(k, "") for k in COLUNAS])
        for c in range(1, len(COLUNAS) + 1):
            ws.cell(ws.max_row, c).font = fonte
            ws.cell(ws.max_row, c).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUNAS))}{ws.max_row}"

    wb.save(PLANILHA_SAIDA)
    print(f"\n✔  {len(resultados)} comunicação(ões) gravadas em: {PLANILHA_SAIDA}")
    print(f"   Período: {periodo[0]} a {periodo[1]}")


# =====================================================================
# MAIN
# =====================================================================

def main():
    ap = argparse.ArgumentParser(description="Monitor DJEN/CNJ")
    ap.add_argument("--dias", type=int, choices=[7, 30, 90], default=JANELA_DIAS,
                    help="Janela de tempo: 7, 30 ou 90 dias (padrão: %(default)s)")
    args = ap.parse_args()

    try:
        resultados, periodo = varrer(args.dias)
    except FileNotFoundError:
        sys.exit(f"ERRO: não encontrei a planilha '{PLANILHA_ENTRADA}'. "
                 f"Coloque-a na mesma pasta ou ajuste PLANILHA_ENTRADA no topo do script.")
    salvar(resultados, periodo)


if __name__ == "__main__":
    main()